import requests
from openpyxl import load_workbook

def buscar_resultados_api():
    response = requests.get('https://api.loterica.digital/resultado.php?jogo=mega-sena')
    resultados_api = response.json()
    return resultados_api

def comparar_resultados(resultados_api, numeros_apostados):
    dezenas_sorteadas = eval(resultados_api[0]['dezenas_sorteadas'])  # Avalia a string como uma lista
    acertos_por_jogo = []

    for numeros_jogo in numeros_apostados:
        acertos = len(set(numeros_jogo) & set(dezenas_sorteadas))
        acertos_por_jogo.append((numeros_jogo, acertos))

    return dezenas_sorteadas, acertos_por_jogo  # Retorna as dezenas sorteadas junto com os acertos por jogo

def ler_dezenas_excel(nome_arquivo):
    workbook = load_workbook(filename=nome_arquivo)
    sheet = workbook.active

    numeros_apostados = []

    for row in sheet.iter_rows(values_only=True):
        numeros_jogo = [cell for cell in row if cell is not None]
        numeros_apostados.append(numeros_jogo)

    return numeros_apostados

def formatar_resultados(dezenas_sorteadas, acertos_por_jogo):
    print("\n### Resultados ###")
    print(f"Números Sorteados: {', '.join(str(num) for num in dezenas_sorteadas)}\n")

    for idx, (numeros_jogo, acertos) in enumerate(acertos_por_jogo, start=1):
        mensagem = ""
        numeros_acertados = ", ".join(str(num) for num in set(numeros_jogo) & set(dezenas_sorteadas))

        if acertos >= 4:
            mensagem = "Parabéns, bilhete premiado!"
            if acertos == 6:
                mensagem = "Parabéns, você é o novo milionário do Brasil!"
        else:
            mensagem = "Bilhete sem premiação."

        print(f"Jogo {idx}: {acertos} número(s) acertado(s) - Números: {numeros_acertados} - {mensagem}")

def main():
    try:
        resultados_api = buscar_resultados_api()
    except requests.RequestException as e:
        print("Erro ao acessar a API:", e)
        return

    if resultados_api[0]['numero_do_jogo'] != 2670:
        print("O jogo ainda não foi sorteado.")
        print("A data do sorteio é dia 31/12/2023 a partir das 21 horas.")
        print("Volte mais tarde para conferir.")
        return

    nome_arquivo = 'apostas.xlsx'  # Altere para o nome do seu arquivo XLSX
    try:
        numeros_apostados = ler_dezenas_excel(nome_arquivo)
    except FileNotFoundError:
        print(f"Arquivo '{nome_arquivo}' não encontrado.")
        return
    except Exception as e:
        print("Erro ao ler o arquivo:", e)
        return

    dezenas_sorteadas, acertos_por_jogo = comparar_resultados(resultados_api, numeros_apostados)
    formatar_resultados(dezenas_sorteadas, acertos_por_jogo)

if __name__ == "__main__":
    main()
